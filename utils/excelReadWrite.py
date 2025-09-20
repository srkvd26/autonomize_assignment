import openpyxl

class ExcelReadWrite:
    @staticmethod
    def write(file_path, products):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products List"
        sheet.append(["Title", "Price", "Rating", "Reviews"])

        for product in products:
            sheet.append([product["title"], product["price"], product["rating"], product["reviews"]])

        workbook.save(file_path)
    
    
    @staticmethod
    def read(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return data